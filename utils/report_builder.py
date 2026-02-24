import io, textwrap
from datetime import datetime
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.lib import colors
from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table,
                                 TableStyle, HRFlowable)
from reportlab.lib.enums import TA_CENTER

# ── Colours ───────────────────────────────────────────────────────────────────
NEON   = '00F5C4'
PURPLE = '7C3AED'
DARK   = '0A0E1A'
CARD   = '111827'
WHITE  = 'FFFFFF'
ALT    = '1a2235'
BORDER = '1e3050'

def _hdr(cell, bg=None, size=10):
    cell.font = Font(bold=True, color=WHITE, size=size, name='Calibri')
    cell.fill = PatternFill('solid', start_color=bg or DARK)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = Border(bottom=Side(style='thin', color=BORDER),
                         right=Side(style='thin', color=BORDER))

def _dat(cell, row):
    cell.font = Font(name='Calibri', size=10, color='D1D5DB')
    cell.alignment = Alignment(vertical='center')
    cell.fill = PatternFill('solid', start_color=CARD if row % 2 else ALT)


def build_excel_report(df_clean, ai_analysis, dax_formulas, cleaning_actions, filename) -> bytes:
    wb = Workbook()

    # ── Sheet 1: Cleaned Data ─────────────────────────────────────────────
    ws = wb.active
    ws.title = "Cleaned Data"
    ws.sheet_properties.tabColor = NEON

    ws.merge_cells(f'A1:{get_column_letter(len(df_clean.columns))}1')
    ws['A1'] = f"✅ Cleaned Data  ·  {df_clean.shape[0]:,} rows × {df_clean.shape[1]} cols  ·  {filename}"
    ws['A1'].font = Font(bold=True, size=12, color=NEON, name='Calibri')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A1'].fill = PatternFill('solid', start_color=DARK)
    ws.row_dimensions[1].height = 26

    for ci, col in enumerate(df_clean.columns, 1):
        c = ws.cell(row=2, column=ci, value=col)
        _hdr(c, bg=CARD)
    ws.row_dimensions[2].height = 22

    for ri, row_data in enumerate(df_clean.itertuples(index=False), 3):
        for ci, val in enumerate(row_data, 1):
            if hasattr(val, 'item'):
                val = val.item()
            elif not isinstance(val, str):
                try:
                    import pandas as _pd
                    if _pd.isna(val):
                        val = None
                except:
                    pass
            c = ws.cell(row=ri, column=ci, value=val)
            _dat(c, ri)

    # Auto width
    for ci, col in enumerate(df_clean.columns, 1):
        mx = max(len(str(col)), df_clean.iloc[:, ci-1].astype(str).str.len().max() if len(df_clean) else 0)
        ws.column_dimensions[get_column_letter(ci)].width = min(max(mx + 2, 12), 42)

    # Totals
    tr = len(df_clean) + 3
    ws.cell(row=tr, column=1, value="TOTALS").font = Font(bold=True, color=NEON, name='Calibri')
    ws.cell(row=tr, column=1).fill = PatternFill('solid', start_color=DARK)
    for ci, col in enumerate(df_clean.columns, 1):
        if df_clean[col].dtype in [np.float64, np.int64, np.float32, np.int32]:
            cl = get_column_letter(ci)
            c = ws.cell(row=tr, column=ci, value=f'=SUM({cl}3:{cl}{tr-1})')
            c.font = Font(bold=True, color=NEON, name='Calibri')
            c.fill = PatternFill('solid', start_color=DARK)
    ws.freeze_panes = 'A3'

    # ── Sheet 2: Summary ─────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary & KPIs")
    ws2.sheet_properties.tabColor = PURPLE
    ws2.column_dimensions['A'].width = 40
    ws2.column_dimensions['B'].width = 55

    ws2.merge_cells('A1:B1')
    ws2['A1'] = "⚡ DataMind AI — Report Summary"
    ws2['A1'].font = Font(bold=True, size=14, color=NEON, name='Calibri')
    ws2['A1'].alignment = Alignment(horizontal='center')
    ws2['A1'].fill = PatternFill('solid', start_color=DARK)
    ws2.row_dimensions[1].height = 30

    r = 3
    def section(ws, title, start_row):
        ws.merge_cells(f'A{start_row}:B{start_row}')
        c = ws.cell(row=start_row, column=1, value=title)
        c.font = Font(bold=True, color=WHITE, size=11, name='Calibri')
        c.fill = PatternFill('solid', start_color=CARD)
        c.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[start_row].height = 22
        return start_row + 1

    r = section(ws2, "  📋 OVERVIEW", r)
    for label, val in [
        ("File", filename),
        ("Rows (cleaned)", f"{df_clean.shape[0]:,}"),
        ("Columns", df_clean.shape[1]),
        ("Domain", ai_analysis.get('domain','')),
        ("Generated", datetime.now().strftime('%Y-%m-%d %H:%M'))
    ]:
        ws2.cell(row=r, column=1, value=label).font = Font(bold=True, name='Calibri', size=10, color='94A3B8')
        ws2.cell(row=r, column=2, value=str(val)).font = Font(name='Calibri', size=10, color='E2E8F0')
        for c in [1,2]: ws2.cell(row=r, column=c).fill = PatternFill('solid', start_color=ALT if r%2 else CARD)
        r += 1
    r += 1

    # Description
    r = section(ws2, "  📝 DATA DESCRIPTION", r)
    ws2.merge_cells(f'A{r}:B{r}')
    c = ws2.cell(row=r, column=1, value=ai_analysis.get('data_description',''))
    c.font = Font(name='Calibri', size=10, color='CBD5E1')
    c.alignment = Alignment(wrap_text=True)
    ws2.row_dimensions[r].height = 55
    r += 2

    def list_section(ws, title, items, start_row):
        row = section(ws, title, start_row)
        for item in items:
            ws.merge_cells(f'A{row}:B{row}')
            c = ws.cell(row=row, column=1, value=f"  • {item}")
            c.font = Font(name='Calibri', size=10, color='94A3B8')
            c.alignment = Alignment(wrap_text=True)
            c.fill = PatternFill('solid', start_color=ALT if row%2 else CARD)
            ws.row_dimensions[row].height = 20
            row += 1
        return row + 1

    r = list_section(ws2, "  🔧 CLEANING ACTIONS", cleaning_actions, r)
    r = list_section(ws2, "  🎯 RECOMMENDED KPIs", ai_analysis.get('kpis',[]), r)
    r = list_section(ws2, "  💡 BUSINESS INSIGHTS", ai_analysis.get('insights',[]), r)

    # Numeric stats
    num_df = df_clean.select_dtypes(include=[np.number])
    if not num_df.empty:
        r = section(ws2, "  📈 NUMERIC STATISTICS", r)
        ws2.column_dimensions['C'].width = 18
        ws2.column_dimensions['D'].width = 18
        ws2.column_dimensions['E'].width = 18
        ws2.column_dimensions['F'].width = 18
        for ci, h in enumerate(['Column','Sum','Average','Min','Max'], 1):
            _hdr(ws2.cell(row=r, column=ci), bg=CARD)
            ws2.cell(row=r, column=ci).value = h
        r += 1
        for col in num_df.columns:
            vals = [col, f"{num_df[col].sum():,.1f}", f"{num_df[col].mean():,.1f}",
                    f"{num_df[col].min():,.1f}", f"{num_df[col].max():,.1f}"]
            for ci, val in enumerate(vals, 1):
                c = ws2.cell(row=r, column=ci, value=val)
                c.font = Font(name='Calibri', size=10, color='E2E8F0')
                c.fill = PatternFill('solid', start_color=ALT if r%2 else CARD)
            r += 1

    # ── Sheet 3: DAX ─────────────────────────────────────────────────────
    ws3 = wb.create_sheet("DAX Formulas")
    ws3.sheet_properties.tabColor = '00B8D9'
    ws3.column_dimensions['A'].width = 35
    ws3.column_dimensions['B'].width = 95

    ws3.merge_cells('A1:B1')
    ws3['A1'] = "📐 DAX Formulas — Power BI Desktop → Modeling → New Measure → Paste"
    ws3['A1'].font = Font(bold=True, size=11, color=NEON, name='Calibri')
    ws3['A1'].alignment = Alignment(horizontal='center')
    ws3['A1'].fill = PatternFill('solid', start_color=DARK)
    ws3.row_dimensions[1].height = 26

    for ci, h in enumerate(['Measure Name', 'DAX Formula'], 1):
        _hdr(ws3.cell(row=2, column=ci), bg=CARD)
        ws3.cell(row=2, column=ci).value = h

    for ri, (name, formula) in enumerate(dax_formulas.items(), 3):
        nc = ws3.cell(row=ri, column=1, value=name)
        nc.font = Font(bold=True, name='Calibri', size=10, color='00F5C4')
        nc.fill = PatternFill('solid', start_color=ALT if ri%2 else CARD)
        fc = ws3.cell(row=ri, column=2, value=formula)
        fc.font = Font(name='Courier New', size=9, color='A5F3FC')
        fc.alignment = Alignment(wrap_text=True)
        fc.fill = PatternFill('solid', start_color='0d1626' if ri%2 else '0a1220')
        ws3.row_dimensions[ri].height = 42

    # ── Sheet 4: Data Dictionary ──────────────────────────────────────────
    ws4 = wb.create_sheet("Data Dictionary")
    ws4.sheet_properties.tabColor = 'F59E0B'
    for ci, w in enumerate([30,15,18,15,18,30], 1):
        ws4.column_dimensions[get_column_letter(ci)].width = w

    ws4.merge_cells('A1:F1')
    ws4['A1'] = "📖 Data Dictionary"
    ws4['A1'].font = Font(bold=True, size=12, color=NEON, name='Calibri')
    ws4['A1'].alignment = Alignment(horizontal='center')
    ws4['A1'].fill = PatternFill('solid', start_color=DARK)

    for ci, h in enumerate(['Column','Type','Non-Null','Nulls','Unique','Sample Values'], 1):
        _hdr(ws4.cell(row=2, column=ci), bg=CARD)
        ws4.cell(row=2, column=ci).value = h

    for ri, col in enumerate(df_clean.columns, 3):
        samples = ', '.join(str(v) for v in df_clean[col].dropna().unique()[:3])
        for ci, val in enumerate([col, str(df_clean[col].dtype), int(df_clean[col].count()),
                                   int(df_clean[col].isnull().sum()), int(df_clean[col].nunique()), samples], 1):
            c = ws4.cell(row=ri, column=ci, value=val)
            c.font = Font(name='Calibri', size=10, color='E2E8F0')
            c.fill = PatternFill('solid', start_color=ALT if ri%2 else CARD)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def build_pdf_report(df_clean, ai_analysis, dax_formulas, cleaning_actions, filename, charts=None) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                             rightMargin=2*cm, leftMargin=2*cm,
                             topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    CNEON  = colors.HexColor('#00F5C4')
    CDARK  = colors.HexColor('#0A0E1A')
    CCARD  = colors.HexColor('#111827')
    CTEXT  = colors.HexColor('#E2E8F0')
    CMUTED = colors.HexColor('#64748b')

    H1 = ParagraphStyle('H1', parent=styles['Heading1'],
         textColor=CNEON, fontSize=24, spaceAfter=4, alignment=TA_CENTER,
         fontName='Helvetica-Bold')
    H2 = ParagraphStyle('H2', parent=styles['Heading2'],
         textColor=CTEXT, fontSize=13, spaceBefore=18, spaceAfter=6,
         fontName='Helvetica-Bold')
    BODY = ParagraphStyle('BODY', parent=styles['Normal'],
           fontSize=10, leading=17, spaceAfter=4, textColor=CMUTED)
    CODE = ParagraphStyle('CODE', parent=styles['Normal'],
           fontSize=8, fontName='Courier', leading=13,
           textColor=colors.HexColor('#A5F3FC'),
           backColor=colors.HexColor('#0d1220'),
           leftIndent=8, rightIndent=8, spaceAfter=4)
    SUB = ParagraphStyle('SUB', parent=styles['Normal'],
          fontSize=10, textColor=CMUTED, alignment=TA_CENTER, spaceAfter=16)

    story = []

    # Cover
    story.append(Spacer(1, 0.8*cm))
    story.append(Paragraph("⚡ DataMind AI — BI Report", H1))
    story.append(Paragraph(
        f"{datetime.now().strftime('%B %d, %Y · %H:%M')}  ·  {filename}  ·  "
        f"{df_clean.shape[0]:,} rows  ·  {ai_analysis.get('domain','General')} domain", SUB))
    story.append(HRFlowable(width='100%', color=CNEON, thickness=1.5))
    story.append(Spacer(1, 0.4*cm))

    def add_section(title, items_or_text, is_list=True):
        story.append(Paragraph(title, H2))
        if is_list:
            for item in items_or_text:
                story.append(Paragraph(f"&bull; &nbsp;{item}", BODY))
        else:
            story.append(Paragraph(items_or_text, BODY))

    # Description
    add_section("Data Overview", ai_analysis.get('data_description',''), is_list=False)

    # Stats table
    story.append(Paragraph("Dataset Statistics", H2))
    tdata = [['Metric', 'Value']]
    for label, val in [
        ("File", filename), ("Domain", ai_analysis.get('domain','')),
        ("Rows (cleaned)", f"{df_clean.shape[0]:,}"),
        ("Columns", str(df_clean.shape[1])),
        ("Numeric columns", str(len(df_clean.select_dtypes(include=[np.number]).columns))),
        ("Text columns", str(len(df_clean.select_dtypes(include='object').columns))),
        ("Missing values", str(df_clean.isnull().sum().sum())),
    ]:
        tdata.append([label, val])
    t = Table(tdata, colWidths=[8*cm, 8*cm], repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), CNEON),
        ('TEXTCOLOR', (0,0), (-1,0), CDARK),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 9),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [CCARD, colors.HexColor('#1a2235')]),
        ('TEXTCOLOR', (0,1), (-1,-1), CTEXT),
        ('GRID', (0,0), (-1,-1), 0.3, colors.HexColor('#1e3050')),
        ('ALIGN', (1,1), (1,-1), 'RIGHT'),
        ('PADDING', (0,0), (-1,-1), 6),
    ]))
    story.append(t)

    add_section("🔧 Cleaning Actions", cleaning_actions)
    add_section("🎯 Recommended KPIs", ai_analysis.get('kpis', []))
    add_section("💡 Business Insights", ai_analysis.get('insights', []))

    # Numeric stats
    num_df = df_clean.select_dtypes(include=[np.number])
    if not num_df.empty:
        story.append(Paragraph("📈 Numeric Column Statistics", H2))
        stat_data = [['Column', 'Sum', 'Average', 'Min', 'Max']]
        for col in num_df.columns:
            stat_data.append([col,
                f"{num_df[col].sum():,.1f}", f"{num_df[col].mean():,.1f}",
                f"{num_df[col].min():,.1f}", f"{num_df[col].max():,.1f}"])
        t2 = Table(stat_data, repeatRows=1)
        t2.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#7C3AED')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,-1), 9),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [CCARD, colors.HexColor('#1a2235')]),
            ('TEXTCOLOR', (0,1), (-1,-1), CTEXT),
            ('GRID', (0,0), (-1,-1), 0.3, colors.HexColor('#1e3050')),
            ('ALIGN', (1,1), (-1,-1), 'RIGHT'),
            ('PADDING', (0,0), (-1,-1), 6),
        ]))
        story.append(t2)

    # DAX
    if dax_formulas:
        story.append(Paragraph("📐 DAX Formulas", H2))
        story.append(Paragraph(
            "Power BI Desktop → Modeling → New Measure → Paste each formula below",
            ParagraphStyle('note', parent=styles['Normal'], fontSize=9, textColor=CMUTED, spaceAfter=10)))
        for name, formula in dax_formulas.items():
            story.append(Paragraph(f"<b><font color='#00F5C4'>{name}</font></b>",
                ParagraphStyle('mn', parent=styles['Normal'], fontSize=10,
                               textColor=CNEON, spaceAfter=3, spaceBefore=10)))
            wrapped = textwrap.fill(formula, 95).replace('&','&amp;').replace('<','&lt;').replace('>','&gt;')
            story.append(Paragraph(wrapped, CODE))

    doc.build(story)
    return buf.getvalue()
