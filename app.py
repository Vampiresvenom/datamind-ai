import streamlit as st
import pandas as pd
import numpy as np
import json, re, io, textwrap, os
from datetime import datetime

# ── Must be first Streamlit call ─────────────────────────────────────────────
st.set_page_config(
    page_title="DataMind AI · BI Agent",
    page_icon="⚡",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ── CSS ──────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Space+Mono:wght@400;700&family=DM+Sans:wght@400;500;600;700&display=swap');
:root{--neon:#00F5C4;--purple:#7C3AED;--bg:#0A0E1A;--card:#111827;--card2:#1a2235;--border:rgba(0,245,196,0.15);--text:#E2E8F0;--muted:#64748b;}
*{font-family:'DM Sans',sans-serif;}
html,body,.stApp{background:var(--bg)!important;color:var(--text)!important;}
#MainMenu,footer,header{visibility:hidden;}
.block-container{padding:2rem 2rem 4rem 2rem!important;max-width:1300px!important;}
[data-testid="stSidebar"]{background:linear-gradient(180deg,#0d1220,#0A0E1A)!important;border-right:1px solid var(--border)!important;}
[data-testid="stSidebar"] *{color:var(--text)!important;}
.stButton>button{background:linear-gradient(135deg,#00F5C4,#7C3AED)!important;color:#0A0E1A!important;border:none!important;border-radius:10px!important;font-weight:700!important;font-size:1rem!important;padding:.7rem 2rem!important;box-shadow:0 0 20px rgba(0,245,196,.25)!important;transition:all .3s!important;}
.stButton>button:hover{transform:translateY(-2px)!important;box-shadow:0 0 35px rgba(0,245,196,.45)!important;}
.stDownloadButton>button{background:var(--card2)!important;color:var(--neon)!important;border:1px solid var(--border)!important;border-radius:10px!important;font-weight:600!important;}
[data-testid="stMetric"]{background:var(--card)!important;border:1px solid var(--border)!important;border-radius:12px!important;padding:1rem!important;}
[data-testid="stMetricValue"]{color:var(--neon)!important;font-weight:700!important;}
.stTabs [data-baseweb="tab-list"]{background:var(--card)!important;border-radius:12px!important;padding:4px!important;border:1px solid var(--border)!important;gap:4px!important;}
.stTabs [data-baseweb="tab"]{border-radius:8px!important;color:var(--muted)!important;font-weight:600!important;}
.stTabs [aria-selected="true"]{background:rgba(0,245,196,.15)!important;color:var(--neon)!important;}
[data-testid="stExpander"]{background:var(--card)!important;border:1px solid var(--border)!important;border-radius:12px!important;}
.stTextInput>div>div>input{background:var(--card)!important;border:1px solid var(--border)!important;border-radius:10px!important;color:var(--text)!important;font-family:'Space Mono',monospace!important;font-size:.85rem!important;}
.stTextInput>div>div>input:focus{border-color:var(--neon)!important;box-shadow:0 0 20px rgba(0,245,196,.2)!important;}
[data-testid="stFileUploader"]{background:var(--card)!important;border:2px dashed var(--border)!important;border-radius:16px!important;}
hr{border-color:var(--border)!important;}
.card{background:var(--card);border:1px solid var(--border);border-radius:14px;padding:1.2rem 1.4rem;margin:.5rem 0;}
.card-neon{border-left:3px solid var(--neon);}
.card-purple{border-left:3px solid var(--purple);}
.card-warn{border-left:3px solid #FBBF24;background:rgba(251,191,36,.05);}
.label{font-family:'Space Mono',monospace;font-size:.72rem;color:var(--neon);letter-spacing:2px;text-transform:uppercase;opacity:.8;margin-bottom:.6rem;}
.dax{background:#0d1220;border:1px solid var(--border);border-radius:10px;padding:1rem;font-family:'Space Mono',monospace;font-size:.82rem;line-height:1.7;color:#a5f3fc;overflow-x:auto;white-space:pre-wrap;}
</style>
""", unsafe_allow_html=True)

# ════════════════════════════════════════════════════════════════
# UTILITY FUNCTIONS (all in one file — no imports needed)
# ════════════════════════════════════════════════════════════════

def build_summary(df):
    lines = [f"Shape: {df.shape[0]} rows x {df.shape[1]} columns\nColumns:"]
    for col in df.columns:
        dtype = str(df[col].dtype)
        nulls = int(df[col].isnull().sum())
        unique = int(df[col].nunique())
        if df[col].dtype in [np.float64,np.int64,np.float32,np.int32]:
            info = f"min={df[col].min():.1f}, max={df[col].max():.1f}, mean={df[col].mean():.1f}"
        else:
            samples = str(df[col].dropna().unique()[:4].tolist())
            info = f"samples={samples}"
        lines.append(f"  {col} | {dtype} | {nulls} nulls | {unique} unique | {info}")
    return "\n".join(lines)


def call_claude(api_key, prompt, max_tokens=2500):
    try:
        import anthropic
        client = anthropic.Anthropic(api_key=api_key)
        r = client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=max_tokens,
            messages=[{"role":"user","content":prompt}]
        )
        text = r.content[0].text.strip()
        return re.sub(r'```[a-z]*\n?','',text).strip()
    except Exception as e:
        raise ValueError(str(e))


def ai_analyze(api_key, filename, summary):
    prompt = f"""You are an expert BI analyst. Analyze this dataset. Reply ONLY with valid JSON — no markdown, no extra text.

File: {filename}
{summary}

Return exactly this structure:
{{
  "data_description": "2-3 sentence summary of what this data is about",
  "domain": "Sales or Finance or HR or Inventory or Marketing or Operations or Other",
  "date_column": "exact column name with dates, or null",
  "numeric_columns": ["list","of","numeric","columns"],
  "category_columns": ["list","of","category","columns"],
  "kpis": ["KPI 1: name — what it measures","KPI 2","KPI 3","KPI 4","KPI 5"],
  "quality_issues": ["issue 1","issue 2"],
  "insights": ["business insight 1","insight 2","insight 3"]
}}"""
    text = call_claude(api_key, prompt, 2000)
    try:
        return json.loads(text)
    except:
        return {
            "data_description":"Dataset analysed successfully.",
            "domain":"General","date_column":None,
            "numeric_columns":[],"category_columns":[],
            "kpis":["Row Count","Total Records"],
            "quality_issues":[],"insights":["Data cleaned and ready for analysis."]
        }


def ai_dax(api_key, filename, df, ai):
    table = re.sub(r'[^a-zA-Z0-9_]','_', filename.replace('.xlsx','').replace('.csv',''))
    cols = {col:{"type":str(df[col].dtype),"sample":[str(v) for v in df[col].dropna().unique()[:3].tolist()]} for col in df.columns}
    prompt = f"""You are a Power BI DAX expert. Write production-ready DAX measures.
Reply ONLY with valid JSON — no markdown, no extra text.

Table name: '{table}'
Domain: {ai.get('domain','General')}
Columns: {json.dumps(cols)}
Date column: {ai.get('date_column','none')}
Numeric columns: {ai.get('numeric_columns',[])}
Category columns: {ai.get('category_columns',[])}

Generate 10-12 measures: totals, averages, count, % of total, YTD (if date), MoM growth (if date), running total (if date), RANKX, max, min, conditional flag.

JSON format only:
{{"Measure Name": "complete DAX formula", "Measure 2": "formula"}}"""
    text = call_claude(api_key, prompt, 3000)
    try:
        return json.loads(text), table
    except:
        return {f"Total (Row Count)": f"Row Count = COUNTROWS('{table}')"}, table


def ai_powerquery(api_key, filename, df_raw, df_clean, actions):
    col_types = {}
    for col in df_clean.columns:
        dtype = str(df_clean[col].dtype)
        col_types[col] = "type number" if ('int' in dtype or 'float' in dtype) else \
                         "type datetime" if 'datetime' in dtype else "type text"
    prompt = f"""Write complete Power Query M code. Return ONLY M code — no markdown, no explanation.

Source file: {filename}
Columns and types: {json.dumps(col_types)}
Cleaning done: {json.dumps(actions)}

Write let...in M script: load Excel, promote headers, remove duplicates, remove empty rows,
trim whitespace, set correct types. Add a comment on each step. User will update file path."""
    text = call_claude(api_key, prompt, 2500)
    if not text.strip().startswith('let'):
        pairs = ", ".join([f'{{"{c}", {t}}}' for c,t in col_types.items()])
        return f"""let
    // Step 1: Load Excel file — update the path below to your actual file
    Source = Excel.Workbook(File.Contents("C:\\\\YourPath\\\\{filename}"), null, true),
    // Step 2: Select first sheet
    Sheet = Source{{0}}[Data],
    // Step 3: Promote first row as column headers
    Headers = Table.PromoteHeaders(Sheet, [PromoteAllScalars=true]),
    // Step 4: Remove duplicate rows
    NoDuplicates = Table.Distinct(Headers),
    // Step 5: Remove fully empty rows
    NoEmpties = Table.SelectRows(NoDuplicates, each not List.IsEmpty(
        List.RemoveMatchingItems(Record.FieldValues(_), {{null, ""}}))),
    // Step 6: Trim whitespace from all text columns
    Trimmed = Table.TransformColumns(NoEmpties, {{}}, Text.Trim),
    // Step 7: Set correct data types for each column
    TypedTable = Table.TransformColumnTypes(Trimmed, {{{pairs}}}),
    Final = TypedTable
in
    Final"""
    return text


def _clean_col(name):
    name = str(name).strip()
    name = re.sub(r'[^a-zA-Z0-9\s_]','',name)
    name = re.sub(r'\s+','_',name).strip('_')
    if name and name[0].isdigit(): name = 'Col_'+name
    return name or 'Column'

def _is_num(s):
    try: float(str(s).replace(',','').replace('$','').replace('%','').strip()); return True
    except: return False

def clean_df(df):
    df_c = df.copy(); actions = []
    old = list(df_c.columns)
    df_c.columns = [_clean_col(c) for c in df_c.columns]
    renamed = [(o,n) for o,n in zip(old,df_c.columns) if o!=n]
    if renamed: actions.append(f"Renamed {len(renamed)} column(s) — removed special characters")
    before = len(df_c)
    df_c.dropna(how='all',inplace=True); df_c.dropna(axis=1,how='all',inplace=True)
    if before-len(df_c)>0: actions.append(f"Removed {before-len(df_c)} fully empty rows")
    dups = df_c.duplicated().sum()
    if dups>0: df_c.drop_duplicates(inplace=True); actions.append(f"Removed {dups} duplicate rows")
    for col in df_c.select_dtypes(include='object').columns:
        df_c[col] = df_c[col].apply(lambda x: x.strip() if isinstance(x,str) else x)
        df_c[col] = df_c[col].replace('',np.nan)
    actions.append("Trimmed whitespace from all text columns")
    converted_num = []
    for col in df_c.select_dtypes(include='object').columns:
        sample = df_c[col].dropna().head(20)
        if len(sample)>0 and sample.apply(_is_num).mean()>0.8:
            try:
                df_c[col] = pd.to_numeric(df_c[col].astype(str).str.replace(',','').str.replace('$','').str.replace('%','').str.strip(),errors='coerce')
                converted_num.append(col)
            except: pass
    if converted_num: actions.append(f"Converted to numeric: {', '.join(converted_num)}")
    converted_dt = []
    for col in df_c.select_dtypes(include='object').columns:
        if any(k in col.lower() for k in ['date','time','year','month','day','dt','created','updated']):
            try:
                result = pd.to_datetime(df_c[col],infer_datetime_format=True,errors='coerce')
                if result.notna().mean()>0.7: df_c[col]=result; converted_dt.append(col)
            except: pass
    if converted_dt: actions.append(f"Parsed date columns: {', '.join(converted_dt)}")
    filled = []
    for col in df_c.select_dtypes(include=[np.number]).columns:
        nulls = df_c[col].isnull().sum()
        if nulls>0: df_c[col]=df_c[col].fillna(df_c[col].median()); filled.append(f"{col}({nulls})")
    if filled: actions.append(f"Filled missing values with median: {', '.join(filled)}")
    df_c.reset_index(drop=True,inplace=True)
    return df_c, actions


PALETTE = ['#00F5C4','#7C3AED','#F59E0B','#F87171','#60A5FA','#34D399','#A78BFA']
LAYOUT = dict(paper_bgcolor='#111827',plot_bgcolor='#111827',
              font=dict(color='#E2E8F0',family='DM Sans'),
              xaxis=dict(gridcolor='rgba(255,255,255,0.05)'),
              yaxis=dict(gridcolor='rgba(255,255,255,0.05)'),
              margin=dict(t=50,l=20,r=20,b=20),
              hoverlabel=dict(bgcolor='#1a2235',bordercolor='#00F5C4'))

def make_charts(df, ai):
    try: import plotly.graph_objects as go; import plotly.express as px
    except: return []
    charts = []
    num_cols = [c for c in (ai.get('numeric_columns') or []) if c in df.columns]
    cat_cols = [c for c in (ai.get('category_columns') or []) if c in df.columns]
    date_col = ai.get('date_column')
    if date_col and date_col not in df.columns: date_col = None
    if not num_cols: num_cols = list(df.select_dtypes(include=[np.number]).columns)
    if not cat_cols: cat_cols = [c for c in df.select_dtypes(include='object').columns if df[c].nunique()<30]

    # Time series
    if date_col and num_cols:
        try:
            cols = num_cols[:2]
            dft = df[[date_col]+cols].dropna().sort_values(date_col)
            fig = go.Figure()
            for i,col in enumerate(cols):
                fig.add_trace(go.Scatter(x=dft[date_col],y=dft[col],name=col,mode='lines+markers',
                    line=dict(color=PALETTE[i],width=2.5,shape='spline'),marker=dict(size=4),
                    fill='tozeroy' if i==0 else 'none',fillcolor='rgba(0,245,196,0.05)' if i==0 else None))
            fig.update_layout(**LAYOUT,title="📈 Trend Over Time",hovermode='x unified',
                legend=dict(orientation='h',yanchor='bottom',y=1.02,bgcolor='rgba(0,0,0,0)'))
            charts.append(fig)
        except: pass

    # Bar chart
    if cat_cols and num_cols:
        try:
            cat,num = cat_cols[0],num_cols[0]
            dfb = df.groupby(cat)[num].sum().reset_index().sort_values(num,ascending=False).head(15)
            fig = go.Figure(go.Bar(x=dfb[cat],y=dfb[num],
                marker=dict(color=dfb[num],colorscale=[[0,'#7C3AED'],[0.5,'#00F5C4'],[1,'#F59E0B']]),
                hovertemplate=f'<b>%{{x}}</b><br>{num}: %{{y:,.0f}}<extra></extra>'))
            fig.update_layout(**LAYOUT,title=f"📊 {num} by {cat}")
            charts.append(fig)
        except: pass

    # Donut
    if cat_cols and num_cols:
        try:
            cat,num = cat_cols[0],num_cols[0]
            if df[cat].nunique()<=12:
                dfp = df.groupby(cat)[num].sum().reset_index()
                fig = go.Figure(go.Pie(labels=dfp[cat],values=dfp[num],hole=0.5,
                    marker=dict(colors=PALETTE,line=dict(color='#0A0E1A',width=2)),
                    hovertemplate='<b>%{label}</b><br>%{value:,.0f} (%{percent})<extra></extra>'))
                fig.update_layout(**LAYOUT,title=f"🥧 {num} Share by {cat}")
                charts.append(fig)
        except: pass

    # Histogram
    if num_cols:
        try:
            col = num_cols[0]
            fig = go.Figure(go.Histogram(x=df[col],nbinsx=30,
                marker=dict(color='#00F5C4',opacity=0.8,line=dict(color='#0A0E1A',width=1))))
            fig.update_layout(**LAYOUT,title=f"📉 Distribution of {col}")
            charts.append(fig)
        except: pass

    # Correlation heatmap
    if len(num_cols)>=3:
        try:
            corr = df[num_cols[:8]].corr()
            fig = go.Figure(go.Heatmap(z=corr.values,x=corr.columns,y=corr.columns,
                colorscale=[[0,'#7C3AED'],[0.5,'#111827'],[1,'#00F5C4']],
                zmin=-1,zmax=1,text=corr.values.round(2),texttemplate='%{text}',
                textfont=dict(size=10),showscale=True))
            fig.update_layout(**LAYOUT,title="🔗 Correlation Matrix")
            charts.append(fig)
        except: pass
    return charts


def make_excel(df_clean, ai, dax, actions, filename):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    wb = Workbook()
    N='00F5C4'; D='0A0E1A'; C='111827'; W='FFFFFF'; A='1a2235'

    def hdr(cell,bg=None):
        cell.font=Font(bold=True,color=W,size=10,name='Calibri')
        cell.fill=PatternFill('solid',start_color=bg or D)
        cell.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
    def dat(cell,row):
        cell.font=Font(name='Calibri',size=10,color='D1D5DB')
        cell.fill=PatternFill('solid',start_color=C if row%2 else A)
        cell.alignment=Alignment(vertical='center')

    # Sheet 1: Data
    ws=wb.active; ws.title="Cleaned Data"
    ncols=len(df_clean.columns)
    ws.merge_cells(f'A1:{get_column_letter(ncols)}1')
    ws['A1']=f"✅ Cleaned — {df_clean.shape[0]:,} rows × {ncols} cols · {filename}"
    ws['A1'].font=Font(bold=True,size=12,color=N,name='Calibri')
    ws['A1'].alignment=Alignment(horizontal='center',vertical='center')
    ws['A1'].fill=PatternFill('solid',start_color=D)
    ws.row_dimensions[1].height=26
    for ci,col in enumerate(df_clean.columns,1):
        c=ws.cell(row=2,column=ci,value=col); hdr(c,bg=C)
    ws.row_dimensions[2].height=22
    for ri,row_data in enumerate(df_clean.itertuples(index=False),3):
        for ci,val in enumerate(row_data,1):
            if hasattr(val,'item'): val=val.item()
            elif not isinstance(val,str):
                try:
                    if pd.isna(val): val=None
                except: pass
            c=ws.cell(row=ri,column=ci,value=val); dat(c,ri)
    for ci,col in enumerate(df_clean.columns,1):
        mx=max(len(str(col)),df_clean.iloc[:,ci-1].astype(str).str.len().max() if len(df_clean) else 0)
        ws.column_dimensions[get_column_letter(ci)].width=min(max(mx+2,12),40)
    tr=len(df_clean)+3
    ws.cell(row=tr,column=1,value="TOTALS").font=Font(bold=True,color=N,name='Calibri')
    for ci,col in enumerate(df_clean.columns,1):
        if str(df_clean[col].dtype) in ['float64','int64','float32','int32']:
            cl=get_column_letter(ci)
            c=ws.cell(row=tr,column=ci,value=f'=SUM({cl}3:{cl}{tr-1})')
            c.font=Font(bold=True,color=N,name='Calibri')
            c.fill=PatternFill('solid',start_color=D)
    ws.freeze_panes='A3'

    # Sheet 2: Summary
    ws2=wb.create_sheet("Summary & KPIs"); ws2.column_dimensions['A'].width=40; ws2.column_dimensions['B'].width=55
    ws2.merge_cells('A1:B1'); ws2['A1']="⚡ DataMind AI — Summary"
    ws2['A1'].font=Font(bold=True,size=13,color=N,name='Calibri')
    ws2['A1'].alignment=Alignment(horizontal='center')
    ws2['A1'].fill=PatternFill('solid',start_color=D); ws2.row_dimensions[1].height=28
    r=3
    def sec(ws,title,row):
        ws.merge_cells(f'A{row}:B{row}')
        c=ws.cell(row=row,column=1,value=f"  {title}")
        c.font=Font(bold=True,color=W,size=11,name='Calibri')
        c.fill=PatternFill('solid',start_color=C)
        c.alignment=Alignment(horizontal='left',vertical='center')
        ws.row_dimensions[row].height=22; return row+1
    def rows_list(ws,items,row):
        for item in items:
            ws.merge_cells(f'A{row}:B{row}')
            c=ws.cell(row=row,column=1,value=f"  • {item}")
            c.font=Font(name='Calibri',size=10,color='94A3B8')
            c.fill=PatternFill('solid',start_color=A if row%2 else C)
            c.alignment=Alignment(wrap_text=True); ws.row_dimensions[row].height=20; row+=1
        return row+1
    r=sec(ws2,"📋 OVERVIEW",r)
    for label,val in [("File",filename),("Rows",f"{df_clean.shape[0]:,}"),("Columns",df_clean.shape[1]),
                       ("Domain",ai.get('domain','')),("Generated",datetime.now().strftime('%Y-%m-%d %H:%M'))]:
        ws2.cell(row=r,column=1,value=label).font=Font(bold=True,name='Calibri',size=10,color='94A3B8')
        ws2.cell(row=r,column=2,value=str(val)).font=Font(name='Calibri',size=10,color='E2E8F0')
        for c in [1,2]: ws2.cell(row=r,column=c).fill=PatternFill('solid',start_color=A if r%2 else C)
        r+=1
    r+=1
    ws2.merge_cells(f'A{r}:B{r}'); c=ws2.cell(row=r,column=1,value=f"  {ai.get('data_description','')}")
    c.font=Font(name='Calibri',size=10,color='CBD5E1'); c.alignment=Alignment(wrap_text=True)
    c.fill=PatternFill('solid',start_color=A); ws2.row_dimensions[r].height=55; r+=2
    r=sec(ws2,"🔧 CLEANING ACTIONS",r); r=rows_list(ws2,actions,r)
    r=sec(ws2,"🎯 KPIs",r); r=rows_list(ws2,ai.get('kpis',[]),r)
    r=sec(ws2,"💡 INSIGHTS",r); r=rows_list(ws2,ai.get('insights',[]),r)

    # Sheet 3: DAX
    ws3=wb.create_sheet("DAX Formulas"); ws3.column_dimensions['A'].width=35; ws3.column_dimensions['B'].width=95
    ws3.merge_cells('A1:B1'); ws3['A1']="📐 DAX Formulas — Modeling → New Measure → Paste"
    ws3['A1'].font=Font(bold=True,size=11,color=N,name='Calibri')
    ws3['A1'].alignment=Alignment(horizontal='center'); ws3['A1'].fill=PatternFill('solid',start_color=D)
    ws3.row_dimensions[1].height=26
    for ci,h in enumerate(['Measure Name','DAX Formula'],1):
        hdr(ws3.cell(row=2,column=ci),bg=C); ws3.cell(row=2,column=ci).value=h
    for ri,(name,formula) in enumerate(dax.items(),3):
        nc=ws3.cell(row=ri,column=1,value=name); nc.font=Font(bold=True,name='Calibri',size=10,color=N)
        nc.fill=PatternFill('solid',start_color=A if ri%2 else C)
        fc=ws3.cell(row=ri,column=2,value=formula); fc.font=Font(name='Courier New',size=9,color='A5F3FC')
        fc.alignment=Alignment(wrap_text=True); fc.fill=PatternFill('solid',start_color='0d1626' if ri%2 else '0a1220')
        ws3.row_dimensions[ri].height=42

    # Sheet 4: Dictionary
    ws4=wb.create_sheet("Data Dictionary")
    for ci,w in enumerate([30,15,18,15,18,30],1): ws4.column_dimensions[get_column_letter(ci)].width=w
    ws4.merge_cells('A1:F1'); ws4['A1']="📖 Data Dictionary"
    ws4['A1'].font=Font(bold=True,size=12,color=N,name='Calibri')
    ws4['A1'].alignment=Alignment(horizontal='center'); ws4['A1'].fill=PatternFill('solid',start_color=D)
    for ci,h in enumerate(['Column','Type','Non-Null','Nulls','Unique','Samples'],1):
        hdr(ws4.cell(row=2,column=ci),bg=C); ws4.cell(row=2,column=ci).value=h
    for ri,col in enumerate(df_clean.columns,3):
        samples=', '.join(str(v) for v in df_clean[col].dropna().unique()[:3])
        for ci,val in enumerate([col,str(df_clean[col].dtype),int(df_clean[col].count()),
                                  int(df_clean[col].isnull().sum()),int(df_clean[col].nunique()),samples],1):
            c=ws4.cell(row=ri,column=ci,value=val); c.font=Font(name='Calibri',size=10,color='E2E8F0')
            c.fill=PatternFill('solid',start_color=A if ri%2 else C)

    out=io.BytesIO(); wb.save(out); return out.getvalue()


def make_pdf(df_clean, ai, dax, actions, filename):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
    from reportlab.lib.enums import TA_CENTER
    buf=io.BytesIO()
    doc=SimpleDocTemplate(buf,pagesize=A4,rightMargin=2*cm,leftMargin=2*cm,topMargin=2*cm,bottomMargin=2*cm)
    styles=getSampleStyleSheet()
    CN=colors.HexColor('#00F5C4'); CD=colors.HexColor('#0A0E1A'); CC=colors.HexColor('#111827')
    CT=colors.HexColor('#E2E8F0'); CM=colors.HexColor('#64748b')
    H1=ParagraphStyle('H1',parent=styles['Heading1'],textColor=CN,fontSize=22,spaceAfter=4,alignment=TA_CENTER,fontName='Helvetica-Bold')
    H2=ParagraphStyle('H2',parent=styles['Heading2'],textColor=CT,fontSize=13,spaceBefore=16,spaceAfter=6,fontName='Helvetica-Bold')
    BODY=ParagraphStyle('B',parent=styles['Normal'],fontSize=10,leading=17,spaceAfter=4,textColor=CM)
    CODE=ParagraphStyle('C',parent=styles['Normal'],fontSize=8,fontName='Courier',leading=13,
                         textColor=colors.HexColor('#A5F3FC'),backColor=colors.HexColor('#0d1220'),
                         leftIndent=8,rightIndent=8,spaceAfter=4)
    SUB=ParagraphStyle('S',parent=styles['Normal'],fontSize=10,textColor=CM,alignment=TA_CENTER,spaceAfter=14)
    story=[]
    story.append(Spacer(1,.6*cm))
    story.append(Paragraph("⚡ DataMind AI — BI Report",H1))
    story.append(Paragraph(f"{datetime.now().strftime('%B %d, %Y · %H:%M')}  ·  {filename}  ·  {df_clean.shape[0]:,} rows  ·  {ai.get('domain','General')}",SUB))
    story.append(HRFlowable(width='100%',color=CN,thickness=1.5))
    story.append(Spacer(1,.3*cm))
    story.append(Paragraph("Data Overview",H2))
    story.append(Paragraph(ai.get('data_description',''),BODY))
    # Stats table
    story.append(Paragraph("Dataset Statistics",H2))
    td=[['Metric','Value']]
    for l,v in [("File",filename),("Domain",ai.get('domain','')),("Rows",f"{df_clean.shape[0]:,}"),
                ("Columns",str(df_clean.shape[1])),("Missing values",str(df_clean.isnull().sum().sum()))]:
        td.append([l,v])
    t=Table(td,colWidths=[8*cm,8*cm],repeatRows=1)
    t.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),CN),('TEXTCOLOR',(0,0),(-1,0),CD),
        ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),('FONTSIZE',(0,0),(-1,-1),9),
        ('ROWBACKGROUNDS',(0,1),(-1,-1),[CC,colors.HexColor('#1a2235')]),
        ('TEXTCOLOR',(0,1),(-1,-1),CT),('GRID',(0,0),(-1,-1),.3,colors.HexColor('#1e3050')),
        ('ALIGN',(1,1),(1,-1),'RIGHT'),('PADDING',(0,0),(-1,-1),6)]))
    story.append(t)
    def add_list(title,items):
        story.append(Paragraph(title,H2))
        for item in items: story.append(Paragraph(f"• {item}",BODY))
    add_list("🔧 Cleaning Actions",actions)
    add_list("🎯 Recommended KPIs",ai.get('kpis',[]))
    add_list("💡 Business Insights",ai.get('insights',[]))
    num_df=df_clean.select_dtypes(include=[np.number])
    if not num_df.empty:
        story.append(Paragraph("📈 Numeric Statistics",H2))
        sd=[['Column','Sum','Average','Min','Max']]
        for col in num_df.columns:
            sd.append([col,f"{num_df[col].sum():,.1f}",f"{num_df[col].mean():,.1f}",
                       f"{num_df[col].min():,.1f}",f"{num_df[col].max():,.1f}"])
        t2=Table(sd,repeatRows=1)
        t2.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),colors.HexColor('#7C3AED')),
            ('TEXTCOLOR',(0,0),(-1,0),colors.white),('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),
            ('FONTSIZE',(0,0),(-1,-1),9),('ROWBACKGROUNDS',(0,1),(-1,-1),[CC,colors.HexColor('#1a2235')]),
            ('TEXTCOLOR',(0,1),(-1,-1),CT),('GRID',(0,0),(-1,-1),.3,colors.HexColor('#1e3050')),
            ('ALIGN',(1,1),(-1,-1),'RIGHT'),('PADDING',(0,0),(-1,-1),6)]))
        story.append(t2)
    if dax:
        story.append(Paragraph("📐 DAX Formulas",H2))
        story.append(Paragraph("Power BI Desktop → Modeling → New Measure → Paste formula",
            ParagraphStyle('n',parent=styles['Normal'],fontSize=9,textColor=CM,spaceAfter=8)))
        for name,formula in dax.items():
            story.append(Paragraph(f"<b><font color='#00F5C4'>{name}</font></b>",
                ParagraphStyle('mn',parent=styles['Normal'],fontSize=10,textColor=CN,spaceAfter=3,spaceBefore=10)))
            wrapped=textwrap.fill(formula,90).replace('&','&amp;').replace('<','&lt;').replace('>','&gt;')
            story.append(Paragraph(wrapped,CODE))
    doc.build(story); return buf.getvalue()


# ════════════════════════════════════════════════════════════════
# SIDEBAR
# ════════════════════════════════════════════════════════════════
with st.sidebar:
    st.markdown("""
    <div style="padding:1.2rem 0 .8rem 0;text-align:center;">
        <div style="font-size:2.2rem;">⚡</div>
        <div style="font-family:'Space Mono',monospace;font-size:1.2rem;font-weight:700;
                    background:linear-gradient(90deg,#00F5C4,#7C3AED);
                    -webkit-background-clip:text;-webkit-text-fill-color:transparent;">DataMind AI</div>
        <div style="font-size:.72rem;color:#475569;letter-spacing:2px;text-transform:uppercase;margin-top:2px;">BI Agent · Free</div>
    </div>
    """, unsafe_allow_html=True)
    st.markdown("<hr>", unsafe_allow_html=True)
    st.markdown('<div class="label">Claude API Key</div>', unsafe_allow_html=True)
    api_key = st.text_input("key", type="password", placeholder="sk-ant-api03-...", label_visibility="collapsed")
    if api_key and len(api_key)>20:
        st.markdown('<div style="color:#00F5C4;font-size:.85rem;margin-top:4px;">✓ Key entered</div>', unsafe_allow_html=True)
    else:
        st.markdown('<div style="font-size:.8rem;color:#475569;margin-top:6px;">Get free key →<br><a href="https://console.anthropic.com" target="_blank" style="color:#00F5C4;">console.anthropic.com</a></div>', unsafe_allow_html=True)
    st.markdown("<hr>", unsafe_allow_html=True)
    st.markdown('<div class="label">Pipeline</div>', unsafe_allow_html=True)
    for n,l in [("01","Upload Excel / CSV"),("02","AI Analyzes Data"),("03","Auto-Clean"),
                ("04","Generate DAX"),("05","Power Query Code"),("06","Build Charts"),("07","Export All")]:
        st.markdown(f'<div style="display:flex;gap:10px;margin-bottom:8px;align-items:center;"><span style="font-family:\'Space Mono\',monospace;font-size:.68rem;color:#00F5C4;opacity:.55;min-width:22px;">{n}</span><span style="font-size:.85rem;color:#94a3b8;">{l}</span></div>', unsafe_allow_html=True)
    st.markdown("<hr>", unsafe_allow_html=True)
    st.markdown('<div style="font-size:.72rem;color:#1e3a5f;text-align:center;">Powered by Claude · Anthropic</div>', unsafe_allow_html=True)

# ════════════════════════════════════════════════════════════════
# HERO
# ════════════════════════════════════════════════════════════════
st.markdown("""
<div style="text-align:center;padding:2rem 0 1.5rem 0;">
    <div style="font-family:'Space Mono',monospace;font-size:.75rem;color:#00F5C4;letter-spacing:3px;
                text-transform:uppercase;margin-bottom:.8rem;opacity:.8;">⚡ Powered by Claude AI</div>
    <h1 style="font-size:clamp(1.8rem,4vw,3rem);font-weight:700;
               background:linear-gradient(135deg,#E2E8F0 0%,#00F5C4 50%,#7C3AED 100%);
               -webkit-background-clip:text;-webkit-text-fill-color:transparent;
               margin:0 0 .8rem 0;line-height:1.15;">
        Your Excel Data,<br>Transformed in Seconds
    </h1>
    <p style="font-size:1.05rem;color:#64748b;max-width:580px;margin:0 auto;line-height:1.7;">
        Upload any Excel or CSV file. Get clean data, DAX formulas,<br>
        Power Query code, interactive charts and a PDF report — instantly.
    </p>
</div>
""", unsafe_allow_html=True)

# Feature pills
c1,c2,c3,c4 = st.columns(4)
for col,(icon,title,desc) in zip([c1,c2,c3,c4],[
    ("🧹","Auto Clean","Removes duplicates, fixes types"),
    ("📐","DAX Formulas","10+ Power BI measures"),
    ("🔄","Power Query","M code for Power BI"),
    ("📊","Dashboard","Charts + PDF report")]):
    with col:
        st.markdown(f"""<div class="card" style="text-align:center;">
        <div style="font-size:1.6rem;margin-bottom:.4rem;">{icon}</div>
        <div style="font-weight:700;font-size:.9rem;color:#E2E8F0;margin-bottom:.2rem;">{title}</div>
        <div style="font-size:.78rem;color:#475569;">{desc}</div></div>""", unsafe_allow_html=True)

st.markdown("<br>", unsafe_allow_html=True)

# API Key gate
if not api_key or len(api_key)<20:
    st.markdown("""<div class="card" style="text-align:center;padding:2rem;">
    <div style="font-size:1.8rem;margin-bottom:.6rem;">🔑</div>
    <div style="font-size:1rem;font-weight:600;color:#E2E8F0;margin-bottom:.4rem;">Enter your Claude API Key in the sidebar</div>
    <div style="font-size:.9rem;color:#64748b;">Free account at <a href="https://console.anthropic.com" target="_blank" style="color:#00F5C4;">console.anthropic.com</a> — $5 free credits ≈ 250 reports</div>
    </div>""", unsafe_allow_html=True)
    st.stop()

# ════════════════════════════════════════════════════════════════
# FILE UPLOAD
# ════════════════════════════════════════════════════════════════
st.markdown('<div class="label">Upload your data</div>', unsafe_allow_html=True)
uploaded = st.file_uploader("files", type=["xlsx","xls","csv"], accept_multiple_files=True, label_visibility="collapsed")

if not uploaded:
    st.markdown('<div style="text-align:center;color:#334155;padding:1.5rem;font-size:.9rem;">↑ Upload one or more Excel / CSV files to begin</div>', unsafe_allow_html=True)
    st.stop()

# Load files
all_dfs = {}
for uf in uploaded:
    try:
        raw = uf.read()
        if uf.name.endswith('.csv'):
            df = pd.read_csv(io.BytesIO(raw))
        else:
            sheets = pd.read_excel(io.BytesIO(raw), sheet_name=None)
            df = list(sheets.values())[0]
        all_dfs[uf.name] = df
    except Exception as e:
        st.error(f"Could not read **{uf.name}**: {e}")

if not all_dfs:
    st.stop()

# Preview
st.markdown("<br>", unsafe_allow_html=True)
st.markdown('<div class="label">Preview</div>', unsafe_allow_html=True)
for fname, df in all_dfs.items():
    with st.expander(f"📄 {fname}  —  {df.shape[0]:,} rows × {df.shape[1]} cols", expanded=True):
        m1,m2,m3,m4 = st.columns(4)
        m1.metric("Rows", f"{df.shape[0]:,}")
        m2.metric("Columns", df.shape[1])
        m3.metric("Missing", f"{df.isnull().sum().sum():,}")
        m4.metric("Duplicates", f"{df.duplicated().sum():,}")
        st.dataframe(df.head(8), use_container_width=True)

# Run button
st.markdown("<br>", unsafe_allow_html=True)
_,cc,_ = st.columns([1,2,1])
with cc:
    run = st.button("⚡  Run Full AI Analysis", use_container_width=True)

if not run:
    st.stop()

# ════════════════════════════════════════════════════════════════
# PIPELINE — runs when button clicked
# ════════════════════════════════════════════════════════════════
for filename, df_raw in all_dfs.items():

    st.markdown(f"""<div class="card card-neon" style="display:flex;align-items:center;gap:12px;margin:1.5rem 0 1rem 0;">
    <span style="font-size:1.2rem;">📄</span>
    <div><div style="font-weight:700;">{filename}</div>
    <div style="color:#64748b;font-size:.85rem;">{df_raw.shape[0]:,} rows · {df_raw.shape[1]} columns</div></div>
    </div>""", unsafe_allow_html=True)

    tab1,tab2,tab3,tab4,tab5 = st.tabs(["🔍 AI Analysis","🧹 Clean Data","📐 DAX Formulas","🔄 Power Query","📊 Dashboard"])

    # ── Step 1: AI Analysis ─────────────────────────────────────────────
    with tab1:
        with st.spinner("🤖 Claude is reading your data..."):
            try:
                summary = build_summary(df_raw)
                ai = ai_analyze(api_key, filename, summary)
                st.session_state[f'{filename}_ai'] = ai
            except Exception as e:
                st.error(f"❌ AI Analysis failed: {e}")
                st.info("Check your API key is correct and has credits remaining.")
                st.stop()

        c1,c2 = st.columns([3,1])
        with c1:
            st.markdown('<div class="label">What is this data?</div>', unsafe_allow_html=True)
            st.markdown(f'<div class="card card-neon"><div style="font-size:.95rem;line-height:1.7;color:#cbd5e1;">{ai.get("data_description","")}</div></div>', unsafe_allow_html=True)
        with c2:
            st.markdown('<div class="label">Domain</div>', unsafe_allow_html=True)
            st.markdown(f'<div class="card" style="text-align:center;"><div style="font-size:.72rem;color:#64748b;text-transform:uppercase;letter-spacing:1px;">Detected as</div><div style="font-size:1.2rem;font-weight:700;color:#00F5C4;margin-top:4px;">{ai.get("domain","General")}</div></div>', unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)
        q,k = st.columns(2)
        with q:
            st.markdown('<div class="label">⚠️ Quality Issues</div>', unsafe_allow_html=True)
            issues = ai.get('quality_issues',[])
            if issues:
                for i in issues: st.markdown(f'<div class="card card-warn" style="font-size:.9rem;">⚠️ {i}</div>', unsafe_allow_html=True)
            else:
                st.markdown('<div class="card card-neon" style="font-size:.9rem;">✓ No major issues found</div>', unsafe_allow_html=True)
        with k:
            st.markdown('<div class="label">🎯 Recommended KPIs</div>', unsafe_allow_html=True)
            for kpi in ai.get('kpis',[]): st.markdown(f'<div class="card card-purple" style="font-size:.88rem;">🎯 {kpi}</div>', unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)
        st.markdown('<div class="label">💡 Business Insights</div>', unsafe_allow_html=True)
        for ins in ai.get('insights',[]): st.markdown(f'<div class="card card-neon" style="font-size:.88rem;">💡 {ins}</div>', unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)
        st.markdown('<div class="label">📋 Column Summary</div>', unsafe_allow_html=True)
        col_df = pd.DataFrame({'Column':df_raw.columns,'Type':df_raw.dtypes.astype(str).values,
                                'Non-Null':df_raw.count().values,'Missing':df_raw.isnull().sum().values,'Unique':df_raw.nunique().values})
        st.dataframe(col_df, use_container_width=True, hide_index=True)

    # ── Step 2: Clean ────────────────────────────────────────────────────
    with tab2:
        with st.spinner("🧹 Cleaning your data..."):
            df_clean, actions = clean_df(df_raw)

        m1,m2,m3,m4 = st.columns(4)
        m1.metric("Rows Before", f"{df_raw.shape[0]:,}")
        m2.metric("Rows After", f"{df_clean.shape[0]:,}")
        m3.metric("Missing Before", f"{df_raw.isnull().sum().sum():,}")
        m4.metric("Missing After", f"{df_clean.isnull().sum().sum():,}")

        st.markdown("<br>", unsafe_allow_html=True)
        st.markdown('<div class="label">✅ Actions Taken</div>', unsafe_allow_html=True)
        for a in actions: st.markdown(f'<div class="card card-neon" style="font-size:.9rem;">✓ {a}</div>', unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)
        st.markdown('<div class="label">Cleaned Data Preview</div>', unsafe_allow_html=True)
        st.dataframe(df_clean, use_container_width=True)

    # ── Step 3: DAX ──────────────────────────────────────────────────────
    with tab3:
        ai = st.session_state.get(f'{filename}_ai', {})
        with st.spinner("📐 Generating DAX formulas..."):
            try:
                dax, table_name = ai_dax(api_key, filename, df_clean, ai)
            except Exception as e:
                st.error(f"DAX generation failed: {e}")
                dax, table_name = {}, filename

        st.markdown(f'<div class="card" style="margin-bottom:1rem;font-size:.9rem;">Power BI Table Name: <code style="color:#00F5C4;">{table_name}</code> &nbsp;·&nbsp; <span style="color:#64748b;">Modeling → New Measure → Paste any formula below</span></div>', unsafe_allow_html=True)

        for name, formula in dax.items():
            with st.expander(f"📐 {name}"):
                st.markdown(f'<div class="dax">{formula}</div>', unsafe_allow_html=True)
                st.code(formula, language="sql")

    # ── Step 4: Power Query ──────────────────────────────────────────────
    with tab4:
        with st.spinner("🔄 Generating Power Query M code..."):
            try:
                pq = ai_powerquery(api_key, filename, df_raw, df_clean, actions)
            except Exception as e:
                st.error(f"Power Query generation failed: {e}")
                pq = "// Generation failed — please try again"

        st.markdown("""<div class="card card-warn" style="margin-bottom:1rem;">
        <div style="color:#FBBF24;font-weight:600;margin-bottom:4px;">How to use in Power BI</div>
        <div style="color:#94a3b8;font-size:.88rem;line-height:1.8;">
        1. Open Power BI Desktop &nbsp;→&nbsp; 2. Home → Transform Data &nbsp;→&nbsp;
        3. View → Advanced Editor &nbsp;→&nbsp; 4. Delete all code → Paste below &nbsp;→&nbsp;
        5. Update file path → Click Done
        </div></div>""", unsafe_allow_html=True)
        st.code(pq, language="javascript")

    # ── Step 5: Charts ───────────────────────────────────────────────────
    with tab5:
        ai = st.session_state.get(f'{filename}_ai', {})
        with st.spinner("📊 Building charts..."):
            charts = make_charts(df_clean, ai)

        if charts:
            for fig in charts:
                st.plotly_chart(fig, use_container_width=True)
        else:
            st.info("Not enough numeric/category data to generate charts automatically.")

    # ── Downloads ─────────────────────────────────────────────────────────
    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown('<div class="label">📥 Download Your Outputs</div>', unsafe_allow_html=True)

    ai  = st.session_state.get(f'{filename}_ai', {})
    base = re.sub(r'[^a-zA-Z0-9_]','_',filename.replace('.xlsx','').replace('.csv',''))
    ts   = datetime.now().strftime('%Y%m%d_%H%M')

    d1,d2,d3,d4 = st.columns(4)
    with d1:
        try:
            xl = make_excel(df_clean, ai, dax, actions, filename)
            st.download_button("📁 Cleaned Excel", data=xl,
                file_name=f"Cleaned_{base}_{ts}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True)
        except Exception as e:
            st.error(f"Excel error: {e}")

    with d2:
        try:
            pdf = make_pdf(df_clean, ai, dax, actions, filename)
            st.download_button("📄 PDF Report", data=pdf,
                file_name=f"Report_{base}_{ts}.pdf",
                mime="application/pdf", use_container_width=True)
        except Exception as e:
            st.error(f"PDF error: {e}")

    with d3:
        dax_txt = f"DAX Formulas for: {filename}\nTable: '{table_name}'\n\n"
        dax_txt += "HOW TO USE: Power BI → Modeling → New Measure → Paste\n\n" + "="*60 + "\n\n"
        for n,f in dax.items(): dax_txt += f"// {n}\n{f}\n\n{'-'*50}\n\n"
        st.download_button("📐 DAX Formulas", data=dax_txt,
            file_name=f"DAX_{base}_{ts}.txt",
            mime="text/plain", use_container_width=True)

    with d4:
        pq_txt = f"Power Query M for: {filename}\n\nHOW TO USE:\nPower BI → Transform Data → Advanced Editor → Paste\n\n" + "="*60 + "\n\n" + pq
        st.download_button("🔄 Power Query", data=pq_txt,
            file_name=f"PowerQuery_{base}_{ts}.txt",
            mime="text/plain", use_container_width=True)

    st.markdown("<hr>", unsafe_allow_html=True)
